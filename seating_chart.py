import openpyxl
import os
import subprocess
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase.pdfmetrics import registerFont

def read_student_names(txt_file):
    """读取座號和姓名对应关系"""
    students = {}
    with open(txt_file, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            if line:
                parts = line.split('\t')
                if len(parts) >= 2:
                    seat_num = parts[0]
                    name = parts[1]
                    students[seat_num] = name
    return students

def check_xelatex():
    """检查 XeLaTeX 是否已安装"""
    try:
        result = subprocess.run(['xelatex', '--version'], capture_output=True, timeout=5)
        return result.returncode == 0
    except (FileNotFoundError, subprocess.TimeoutExpired):
        return False

def generate_latex(excel_file, txt_file, output_pdf):
    """使用 LaTeX 生成 PDF"""
    print("使用 LaTeX 生成 PDF（推荐方案）...")
    
    # 读取学生姓名映射
    students = read_student_names(txt_file)
    
    # 读取Excel座位表
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 转换为数据结构
    seating_layout = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value if cell.value is not None else None)
        seating_layout.append(row_data)
    
    return generate_latex_from_data(seating_layout, students, output_pdf)


def generate_latex_from_data(seating_layout, students, output_pdf):
    """使用 LaTeX 生成 PDF（直接从数据）"""
    print("使用 LaTeX 生成 PDF（推荐方案）...")
    
    # 生成 LaTeX 源码
    latex_content = r"""\documentclass[12pt,a4paper]{article}
\usepackage{ctex}  % 支持中文
\usepackage{geometry}
\usepackage{array}
\usepackage{colortbl}
\usepackage{xcolor}

% 设置页面边距
\geometry{left=1.5cm,right=1.5cm,top=2cm,bottom=2cm}

% 自定义颜色
\definecolor{seatedge}{RGB}{100,100,100}
\definecolor{seatnumber}{RGB}{102,126,234}
\definecolor{podiumbg}{RGB}{102,126,234}
\definecolor{podiumtext}{RGB}{255,255,255}

% 自定义列样式
\newcolumntype{S}{>{\centering\arraybackslash}m{1.5cm}}
\newcolumntype{M}{>{\centering\arraybackslash}m{2.5cm}}

\begin{document}

\begin{center}
"""

    # 生成座位表表格（老师视角：上下颠倒+左右镜像）
    latex_content += r"\begin{tabular}{"
    
    # 确定列数
    max_cols = max(len(row) for row in seating_layout) if seating_layout else 12
    for _ in range(max_cols):
        latex_content += "S"
    latex_content += "}\n"
    
    # 读取所有行并反转
    for row in reversed(seating_layout):
        # 反转列顺序
        row_cells = []
        for seat_num in reversed(row):
            if seat_num is not None:
                seat_num_str = str(seat_num).strip()
                if seat_num_str in students:
                    name = students[seat_num_str]
                    # 姓名竖排显示
                    vertical_name = '\\ '.join(name)
                    row_cells.append(f"\\colorbox{{white}}{{\\parbox{{1.3cm}}{{\\centering \\textcolor{{seatnumber}}{{\\textbf{{{seat_num_str}}}}}\\\\ \\color{{black}}\\small {vertical_name}}}}}")
                else:
                    row_cells.append(f"\\colorbox{{white}}{{\\parbox{{1.3cm}}{{\\centering \\textcolor{{seatnumber}}{{\\textbf{{{seat_num_str}}}}}}}}}")
            else:
                # 空位显示空白
                row_cells.append("\\colorbox{white}{\\phantom{x}}")
        
        latex_content += " & ".join(row_cells) + " \\\\\n"
    
    latex_content += r"\end{tabular}" + "\n"
    
    # 添加讲台
    latex_content += r"""
\vspace{0.5cm}
\colorbox{podiumbg}{\parbox{10cm}{\centering \color{podiumtext}\large \textbf{講台}}}
"""
    
    latex_content += r"""
\end{center}
\end{document}
"""
    
    # 保存 LaTeX 文件
    latex_file = 'seating_chart.tex'
    with open(latex_file, 'w', encoding='utf-8') as f:
        f.write(latex_content)
    
    print(f"LaTeX 源码已生成: {latex_file}")
    
    # 使用 XeLaTeX 编译 PDF
    try:
        print("正在使用 XeLaTeX 编译 PDF...")
        subprocess.run(['xelatex', '-interaction=nonstopmode', '-output-directory', '.', latex_file], 
                      check=True, capture_output=True)
        
        # 编译两次以正确处理交叉引用
        subprocess.run(['xelatex', '-interaction=nonstopmode', '-output-directory', '.', latex_file], 
                      check=True, capture_output=True)
        
        print(f"PDF 已成功生成: {output_pdf}")
        
        # 清理临时文件
        temp_files = ['seating_chart.aux', 'seating_chart.log']
        for temp_file in temp_files:
            if os.path.exists(temp_file):
                os.remove(temp_file)
                
        return True
                
    except subprocess.CalledProcessError as e:
        print(f"XeLaTeX 编译失败: {e}")
        return False
    except FileNotFoundError:
        print("错误: 未找到 XeLaTeX 命令")
        return False

def generate_reportlab(excel_file, txt_file, output_pdf):
    """使用 ReportLab 生成 PDF（备用方案）"""
    print("使用 ReportLab 生成 PDF（备用方案）...")
    
    # 读取学生姓名映射
    students = read_student_names(txt_file)
    
    # 读取Excel座位表
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active
    
    # 转换为数据结构
    seating_layout = []
    for row in ws.iter_rows():
        row_data = []
        for cell in row:
            row_data.append(cell.value if cell.value is not None else None)
        seating_layout.append(row_data)
    
    return generate_reportlab_from_data(seating_layout, students, output_pdf)


def generate_reportlab_from_data(seating_layout, students, output_pdf):
    """使用 ReportLab 生成 PDF（直接从数据）"""
    print("使用 ReportLab 生成 PDF（备用方案）...")
    
    # 创建PDF（使用直向A4）
    c = canvas.Canvas(output_pdf, pagesize=A4)
    width, height = A4
    
    # 设置字体（優先使用 Windows 系統字體，支援繁體中文）
    font_name = None
    base_dir = os.path.dirname(os.path.abspath(__file__))

    # 嘗試使用 Windows 系統字體
    font_paths = [
        ('MicrosoftJhengHei', 'C:\\Windows\\Fonts\\msjh.ttc'),  # 微軟正黑體
        ('MicrosoftYaHei', 'C:\\Windows\\Fonts\\msyh.ttc'),  # 微軟雅黑
        ('KaiTi', 'C:\\Windows\\Fonts\\simkai.ttf'),  # 楷體
        ('SimSun', 'C:\\Windows\\Fonts\\simsun.ttc'),  # 宋體
    ]
    
    for name, path in font_paths:
        try:
            if os.path.exists(path):
                pdfmetrics.registerFont(TTFont(name, path))
                font_name = name
                print(f"成功加载字体: {name}")
                break
        except Exception as e:
            print(f"加载字体失败 {name}: {e}")
            continue
    
    if font_name is None:
        print("警告: 无法加载中文字体，将使用默认字体（中文可能无法显示）")
        font_name = 'Helvetica'
    
    # 计算座位格大小（适应直向A4，12列，竖排姓名）
    cell_width = 1.5 * cm  # 增加寬度
    cell_height = 2.5 * cm  # 增加高度
    margin_left = 1.5 * cm  # 減少左邊距
    margin_top = 2.5 * cm  # 減少上邊距
    
    # 遍历座位表数据（老师视角：上下颠倒+左右镜像）
    # 反转行顺序：后排在上方，前排在下方
    for row_idx, row in enumerate(reversed(seating_layout)):
        y = margin_top + row_idx * cell_height
        
        # 反转列顺序：从右往左，左右镜像
        for col_idx, seat_num in enumerate(reversed(row)):
            x = margin_left + col_idx * cell_width
            
            # 检查单元格是否有值
            if seat_num is not None:
                seat_num_str = str(seat_num).strip()
                
                # 绘制白色背景
                c.setFillColorRGB(1, 1, 1)  # 白色背景
                c.rect(x, height - y - cell_height, cell_width, cell_height, fill=True, stroke=False)
                
                # 绘制座位格边框
                c.setStrokeColorRGB(0, 0, 0)  # 黑色边框
                c.setLineWidth(2)
                c.rect(x, height - y - cell_height, cell_width, cell_height, fill=False, stroke=True)
                
                # 绘制中间分隔线（座號在上，姓名在下）
                mid_y = height - y - cell_height * 0.35  # 座號占35%高度
                c.line(x, mid_y, x + cell_width, mid_y)
                
                # 绘制座號（上方格子，稍微偏下）
                c.setFillColorRGB(0, 0, 0)
                c.setFont(font_name, 11)
                c.drawCentredString(x + cell_width/2, height - y - cell_height * 0.22, seat_num_str)  # 稍微偏下

                # 绘制姓名（下方格子，竖排）
                c.setFillColorRGB(0, 0, 0)
                c.setFont(font_name, 12)
                if seat_num_str in students:
                    name = students[seat_num_str]
                    # 真正的中文竖排：每个字从上到下排列
                    char_spacing = 0.45*cm  # 字间距
                    # 从下方格子的顶部开始
                    start_y = mid_y - 0.4*cm
                    for i, char in enumerate(name):
                        y_pos = start_y - i * char_spacing
                        c.drawCentredString(x + cell_width/2, y_pos, char)
    
    # 绘制讲台格子（在最下面，老师视角）
    podium_width = 3 * cell_width  # 跨越中间3列
    podium_height = 1.0 * cm  # 增加高度
    podium_x = margin_left + 4.5 * cell_width  # 从中间开始
    # 计算座位表的总行数，把讲台放在最下方
    total_rows = len(seating_layout)
    podium_y = margin_top + total_rows * cell_height + 0.3 * cm  # 在座位表下方
    
    # 绘制讲台背景
    c.setFillColorRGB(0.9, 0.9, 0.9)  # 浅灰色背景
    c.rect(podium_x, height - podium_y - podium_height, podium_width, podium_height, fill=True, stroke=False)
    
    # 绘制讲台边框
    c.setStrokeColorRGB(0, 0, 0)
    c.setLineWidth(2)
    c.rect(podium_x, height - podium_y - podium_height, podium_width, podium_height, fill=False, stroke=True)
    
    # 绘制"講台"文字（横向，居中）
    c.setFillColorRGB(0, 0, 0)
    c.setFont(font_name, 14)  # 增大字体
    c.drawCentredString(podium_x + podium_width/2, height - podium_y - podium_height/2 - 0.05*cm, "講台")
    
    # 保存PDF
    c.save()
    print(f"PDF已生成: {output_pdf}")
    return True

if __name__ == "__main__":
    excel_file = "座位表.xlsx"
    txt_file = "座號對應名字.txt"
    output_pdf = "座位表.pdf"
    
    print("=" * 50)
    print("座位表 PDF 生成器")
    print("=" * 50)
    
    # 检查是否安装了 XeLaTeX
    has_xelatex = check_xelatex()
    
    if has_xelatex:
        print("\n✓ 检测到 XeLaTeX 已安装，使用 LaTeX 生成 PDF（推荐）")
        success = generate_latex(excel_file, txt_file, output_pdf)
    else:
        print("\n✗ 未检测到 XeLaTeX，使用 ReportLab 生成 PDF（备用方案）")
        print("\n提示：安装 XeLaTeX 可以获得更好的 PDF 品质")
        print("安装命令（需要 Chocolatey）: choco install texlive")
        print("或访问: https://www.tug.org/texlive/")
        print()
        success = generate_reportlab(excel_file, txt_file, output_pdf)
    
    if success:
        print(f"\n✓ PDF 生成成功！")
        print(f"  文件位置: {os.path.abspath(output_pdf)}")
    else:
        print(f"\n✗ PDF 生成失败")